"""Broward / Palm Beach native-export import contract.

Stages: A file-audit (no writes) → B parse → C identity → D dry-run → E load.
Synthetic fixtures must be marked TEST_ONLY. Never treat them as agency data.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Iterable

PARSER_VERSION = "enhanced-county-import-v1"

OCC_PREFIX = re.compile(r"^([A-Z]{2,6})(\d{4,})$")
AGENCY_PHONES = {
    "9547654400",
    "9548314000",
    "5612335525",
    "5612335000",
    "5612335001",
    "311",
}

PERMIT_ALIASES = {
    "permit_record_id": ["permit_record_id", "record_id", "permitid", "permit_id", "objectid"],
    "permit_number": ["permit_number", "permitno", "permit_no", "permit#", "permit_num", "permit"],
    "jurisdiction": ["jurisdiction", "ahj", "issuing_jurisdiction", "source_jurisdiction"],
    "municipality": ["municipality", "city", "muni", "city_name"],
    "property_address": ["property_address", "site_address", "address", "street_address", "location"],
    "parcel_id": ["parcel_id", "folio", "pcn", "parcel", "folio_id", "property_id"],
    "permit_type_raw": ["permit_type", "permittype", "type", "permit_subtype", "work_type"],
    "work_description": ["work_description", "description", "job_description", "scope", "comments"],
    "contractor_name_raw": ["contractor_name", "contractor", "license_holder", "qualifier"],
    "contractor_firm_raw": ["contractor_firm", "company", "firm", "business_name", "company_name"],
    "contractor_license_raw": [
        "contractor_license",
        "license_number",
        "dbpr_license",
        "full_license",
        "license_no",
        "state_license",
    ],
    "local_contractor_id": [
        "local_contractor_id",
        "cc_number",
        "cc_no",
        "county_contractor_id",
        "contractor_id",
        "seven_digit_id",
        "contractorid",
    ],
    "application_date": ["application_date", "applied", "app_date", "date_applied", "submit_date"],
    "issue_date": ["issue_date", "issued", "date_issued", "issuedate"],
    "expiration_date": ["expiration_date", "expires", "expire_date", "exp_date"],
    "final_date": ["final_date", "close_date", "closed", "finaled", "co_date"],
    "status_raw": ["status", "permit_status", "current_status", "status_code"],
    "valuation": ["valuation", "job_value", "declared_value", "jobvalue", "est_value", "value"],
    "owner_builder": ["owner_builder", "ownerbuilder", "owner_builder_flag"],
    "source_updated_at": ["last_updated", "updated_at", "modified", "source_updated_at"],
}

CERT_ALIASES = {
    "local_credential_key": [
        "local_credential_key",
        "certificate_id",
        "cert_id",
        "cc_number",
        "credential_id",
        "license_id",
        "contractor_id",
    ],
    "certificate_number_raw": ["certificate_number", "cert_no", "credential_number", "license_number"],
    "classification_raw": ["classification", "class_code", "trade", "category", "class_description"],
    "credential_type": ["credential_type", "type", "license_type", "record_type"],
    "person_name_raw": ["person_name", "contractor_name", "name", "license_holder", "qualifier"],
    "firm_name_raw": ["firm_name", "company", "business_name", "company_name"],
    "dbpr_license_raw": ["dbpr_license", "state_license", "full_license", "dbpr_number"],
    "status_raw": ["status", "license_status", "current_status"],
    "issue_date": ["issue_date", "issued", "date_issued"],
    "renewal_date": ["renewal_date", "renewed"],
    "expiration_date": ["expiration_date", "expires", "exp_date"],
    "phone": ["phone", "phone1", "business_phone", "telephone"],
    "phone2": ["phone2", "additional_phone", "alt_phone"],
    "email": ["email", "email1", "business_email"],
    "email2": ["email2", "additional_email", "alt_email"],
    "website": ["website", "url", "web"],
    "mailing_address": ["mailing_address", "mail_address"],
    "physical_address": ["physical_address", "business_address", "address"],
    "insurance_status_raw": ["insurance_status", "liability_status"],
    "insurance_expiration": ["insurance_expiration", "liability_exp"],
    "workers_comp_status_raw": ["workers_comp_status", "wc_status"],
    "bond_status_raw": ["bond_status"],
    "bond_amount": ["bond_amount"],
}


def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (h or "").strip().lower()).strip("_")


def normalize_full_license(raw: str | None) -> str:
    if not raw:
        return ""
    return re.sub(r"[\s\-_.]", "", str(raw)).upper()


def numeric_core_only(raw: str | None) -> bool:
    n = normalize_full_license(raw)
    if not n:
        return False
    if OCC_PREFIX.match(n):
        return False
    return bool(re.fullmatch(r"\d{4,}", n))


def has_occupation_prefix(raw: str | None) -> bool:
    return bool(OCC_PREFIX.match(normalize_full_license(raw)))


def parse_money(raw: Any) -> tuple[str | None, float | None]:
    if raw is None:
        return None, None
    s = str(raw).strip()
    if not s or s in {".", "-", "NA", "N/A", "null"}:
        return None, None
    raw_s = s
    cleaned = re.sub(r"[,$]", "", s)
    try:
        val = float(cleaned)
    except ValueError:
        return raw_s, None
    return raw_s, val


def is_agency_phone(raw: str | None) -> bool:
    digits = re.sub(r"\D", "", raw or "")
    return digits in AGENCY_PHONES or digits[-10:] in AGENCY_PHONES


def map_aliases(columns: list[str], aliases: dict[str, list[str]]) -> dict[str, str | None]:
    by_norm = {norm_header(c): c for c in columns}
    out: dict[str, str | None] = {}
    for canonical, names in aliases.items():
        hit = None
        for n in names:
            if norm_header(n) in by_norm:
                hit = by_norm[norm_header(n)]
                break
        out[canonical] = hit
    return out


def classify_identity(
    dbpr: str | None,
    dbpr_exists: bool,
    local_id: str | None,
    local_crosswalk: bool,
    name_only: bool,
    ambiguous: bool,
) -> tuple[str, str]:
    full = normalize_full_license(dbpr)
    if full and numeric_core_only(full):
        return "UNRESOLVED", "UNRESOLVED"
    if full and has_occupation_prefix(full) and dbpr_exists:
        return "CONFIRMED", "FULL_DBPR_LICENSE"
    if local_id and local_crosswalk and dbpr_exists:
        return "CONFIRMED", "LOCAL_LICENSE_CROSSWALK"
    if ambiguous:
        return "UNRESOLVED", "REVIEW_REQUIRED"
    if name_only:
        return "REVIEW_REQUIRED", "REVIEW_REQUIRED"
    return "UNRESOLVED", "UNRESOLVED"


def currentness_from_raw(status_raw: str | None, credential_type: str | None) -> str:
    t = (credential_type or "").lower()
    s = (status_raw or "").lower()
    if "installer" in t:
        return "INSTALLER_REGISTRATION"
    if "enroll" in t or "state enroll" in t:
        return "STATE_ENROLLED"
    if "preempt" in s or "preempt" in t:
        return "PREEMPTED_CLASS"
    if "revok" in s:
        return "REVOKED"
    if "expir" in s or "lapsed" in s:
        return "EXPIRED"
    if "historic" in s or "inactive" in s:
        return "HISTORICAL_LOCAL_LICENSE"
    if "active" in s or "current" in s:
        return "CURRENT_LOCAL_AUTHORIZATION"
    return "UNKNOWN"


def permit_record_key(source_system: str, jurisdiction: str, permit_number: str, source_id: str | None) -> str:
    if source_id:
        return f"{source_system}|id|{source_id}"
    return f"{source_system}|{jurisdiction}|{permit_number}"


def read_tabular(path: Path, data: bytes) -> tuple[list[str], list[dict[str, str]], list[str]]:
    """Return columns, rows, sheet_names. CSV/TSV/JSON/XLSX."""
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        text = data.decode("utf-8-sig", errors="replace")
        dialect = csv.excel_tab if suffix == ".tsv" or "\t" in text.splitlines()[:2][-1:] else csv.excel
        reader = csv.DictReader(StringIO(text), dialect=dialect)
        cols = reader.fieldnames or []
        rows = [{k: (v or "").strip() for k, v in r.items() if k} for r in reader]
        return list(cols), rows, [path.name]
    if suffix == ".json":
        payload = json.loads(data.decode("utf-8"))
        if isinstance(payload, dict) and "rows" in payload:
            payload = payload["rows"]
        if not isinstance(payload, list) or not payload:
            return [], [], [path.name]
        cols = list(payload[0].keys())
        rows = [{str(k): "" if v is None else str(v) for k, v in r.items()} for r in payload]
        return cols, rows, [path.name]
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        ws = wb[sheets[0]]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            return [], [], sheets
        cols = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(header)]
        rows = []
        for tup in it:
            if tup is None or all(v is None or str(v).strip() == "" for v in tup):
                continue
            rows.append({cols[i]: "" if i >= len(tup) or tup[i] is None else str(tup[i]).strip() for i in range(len(cols))})
        return cols, rows, sheets
    raise ValueError(f"unsupported format: {suffix}")


SOURCES = {
    "broward-permits": {
        "county": "broward",
        "agency": "Broward County Building Code Division",
        "source_system": "broward_bcs_permits",
        "kind": "permit",
        "aliases": PERMIT_ALIASES,
        "required": ["permit_number"],
    },
    "broward-certs": {
        "county": "broward",
        "agency": "Broward County Building Code Division / CEB",
        "source_system": "broward_bcs_contractor",
        "kind": "cert",
        "aliases": CERT_ALIASES,
        "required": ["local_credential_key"],
    },
    "pbc-permits": {
        "county": "palm-beach",
        "agency": "Palm Beach County PZB Building Division",
        "source_system": "pbc_epzb_permits",
        "kind": "permit",
        "aliases": PERMIT_ALIASES,
        "required": ["permit_number"],
    },
    "pbc-certs": {
        "county": "palm-beach",
        "agency": "Palm Beach County PZB Contractor Regulations",
        "source_system": "pbc_contractor_certs",
        "kind": "cert",
        "aliases": CERT_ALIASES,
        "required": ["local_credential_key"],
    },
}


@dataclass
class FileAudit:
    filename: str
    sha256: str
    file_format: str
    sheet_names: list[str]
    row_count: int
    column_names: list[str]
    inferred_mappings: dict[str, str | None]
    unmapped_fields: list[str]
    parser_version: str = PARSER_VERSION
    test_only: bool = False


def stage_a(path: Path, source: str) -> tuple[FileAudit, list[dict[str, str]]]:
    cfg = SOURCES[source]
    data = path.read_bytes()
    cols, rows, sheets = read_tabular(path, data)
    mapping = map_aliases(cols, cfg["aliases"])
    mapped_src = {v for v in mapping.values() if v}
    unmapped = [c for c in cols if c not in mapped_src]
    audit = FileAudit(
        filename=path.name,
        sha256=sha256_bytes(data),
        file_format=path.suffix.lower().lstrip("."),
        sheet_names=sheets,
        row_count=len(rows),
        column_names=cols,
        inferred_mappings=mapping,
        unmapped_fields=unmapped,
        test_only=path.name.upper().startswith("TEST_ONLY"),
    )
    return audit, rows


def _get(row: dict[str, str], mapping: dict[str, str | None], key: str) -> str:
    col = mapping.get(key)
    if not col:
        return ""
    return (row.get(col) or "").strip()


def stage_b_permits(rows: list[dict[str, str]], mapping: dict[str, str | None], source_system: str) -> list[dict[str, Any]]:
    out = []
    for row in rows:
        number = _get(row, mapping, "permit_number")
        jur = _get(row, mapping, "jurisdiction") or "unknown"
        sid = _get(row, mapping, "permit_record_id") or None
        raw_val, val = parse_money(_get(row, mapping, "valuation"))
        lic = _get(row, mapping, "contractor_license_raw")
        local = _get(row, mapping, "local_contractor_id")
        name = _get(row, mapping, "contractor_name_raw") or _get(row, mapping, "contractor_firm_raw")
        parsed = {
            "source_system": source_system,
            "source_jurisdiction": jur,
            "municipality": _get(row, mapping, "municipality") or None,
            "permit_number": number,
            "source_record_id": sid,
            "record_key": permit_record_key(source_system, jur, number, sid) if number else None,
            "permit_type_raw": _get(row, mapping, "permit_type_raw") or None,
            "work_description": _get(row, mapping, "work_description") or None,
            "property_address": _get(row, mapping, "property_address") or None,
            "parcel_id": _get(row, mapping, "parcel_id") or None,
            "contractor_name_raw": name or None,
            "contractor_license_raw": lic or None,
            "contractor_license_normalized": normalize_full_license(lic) or None,
            "local_contractor_id": local or None,
            "application_date": _get(row, mapping, "application_date") or None,
            "issue_date": _get(row, mapping, "issue_date") or None,
            "expiration_date": _get(row, mapping, "expiration_date") or None,
            "final_date": _get(row, mapping, "final_date") or None,
            "status_raw": _get(row, mapping, "status_raw") or "unknown",
            "status_normalized": "unknown",
            "valuation_raw": raw_val,
            "valuation": val,
            "raw_payload": row,
        }
        out.append(parsed)
    return out


def stage_c_identity(parsed: list[dict[str, Any]], known_dbpr: set[str] | None = None) -> list[dict[str, Any]]:
    known = known_dbpr or set()
    for p in parsed:
        lic = p.get("contractor_license_normalized") or normalize_full_license(
            p.get("contractor_license_raw")
        )
        local = p.get("local_contractor_id")
        name = p.get("contractor_name_raw")
        exists = lic in known if known else has_occupation_prefix(lic)
        name_only = bool(name) and not lic and not local
        ambiguous = bool(name) and len(str(name).split()) < 2
        state, method = classify_identity(lic, exists, local, False, name_only, ambiguous)
        p["identity_state"] = state
        p["identity_method"] = method
        p["issued_activity_eligible"] = bool(p.get("issue_date"))
    return parsed


def quality_report(parsed: list[dict[str, Any]], audit: FileAudit) -> dict[str, Any]:
    keys = [p.get("record_key") for p in parsed if p.get("record_key")]
    dup = len(keys) - len(set(keys))
    statuses = sorted({p.get("status_raw") or "" for p in parsed})
    return {
        "filename": audit.filename,
        "sha256": audit.sha256,
        "test_only": audit.test_only,
        "raw_rows": audit.row_count,
        "parsed_rows": len(parsed),
        "unique_records": len(set(keys)),
        "duplicates": dup,
        "missing_ids": sum(1 for p in parsed if not p.get("record_key")),
        "full_dbpr_licenses": sum(1 for p in parsed if has_occupation_prefix(p.get("contractor_license_raw"))),
        "local_licenses": sum(1 for p in parsed if p.get("local_contractor_id")),
        "name_only_records": sum(1 for p in parsed if p.get("identity_method") == "REVIEW_REQUIRED"),
        "CONFIRMED": sum(1 for p in parsed if p.get("identity_state") == "CONFIRMED"),
        "HIGH_CONFIDENCE": sum(1 for p in parsed if p.get("identity_state") == "HIGH_CONFIDENCE"),
        "REVIEW_REQUIRED": sum(1 for p in parsed if p.get("identity_state") == "REVIEW_REQUIRED"),
        "UNRESOLVED": sum(1 for p in parsed if p.get("identity_state") == "UNRESOLVED"),
        "date_coverage_issue_date": sum(1 for p in parsed if p.get("issue_date")),
        "issued_activity_eligible": sum(1 for p in parsed if p.get("issued_activity_eligible")),
        "valuation_coverage": sum(1 for p in parsed if p.get("valuation") is not None),
        "missing_valuation_not_zero": sum(1 for p in parsed if p.get("valuation") is None),
        "status_values": statuses,
        "email_coverage": 0,
        "phone_coverage": 0,
        "unmapped_fields": audit.unmapped_fields,
        "parser_version": PARSER_VERSION,
    }


def discovery_report(audit: FileAudit, parsed: list[dict[str, Any]]) -> dict[str, Any]:
    q = quality_report(parsed, audit)
    q["column_names"] = audit.column_names
    q["sheet_names"] = audit.sheet_names
    q["inferred_field_mappings"] = audit.inferred_mappings
    q["license_format_distribution"] = {
        "occupation_prefixed": q["full_dbpr_licenses"],
        "numeric_core_only": sum(1 for p in parsed if numeric_core_only(p.get("contractor_license_raw"))),
        "blank": sum(1 for p in parsed if not p.get("contractor_license_raw")),
    }
    q["duplicate_key_rate"] = (q["duplicates"] / q["parsed_rows"]) if q["parsed_rows"] else 0
    q["missing_identity_field_rate"] = (q["missing_ids"] / q["parsed_rows"]) if q["parsed_rows"] else 0
    return q


def run_discovery(path: Path, source: str) -> dict[str, Any]:
    audit, rows = stage_a(path, source)
    cfg = SOURCES[source]
    if cfg["kind"] == "permit":
        parsed = stage_b_permits(rows, audit.inferred_mappings, cfg["source_system"])
        parsed = stage_c_identity(parsed)
    else:
        parsed = []
        for row in rows:
            key = ""
            for canon in ("local_credential_key", "certificate_number_raw"):
                col = audit.inferred_mappings.get(canon)
                if col:
                    key = (row.get(col) or "").strip()
                    if key:
                        break
            parsed.append(
                {
                    "record_key": key or None,
                    "contractor_license_raw": (row.get(audit.inferred_mappings.get("dbpr_license_raw") or "") or ""),
                    "local_contractor_id": key,
                    "status_raw": (row.get(audit.inferred_mappings.get("status_raw") or "") or ""),
                    "issue_date": (row.get(audit.inferred_mappings.get("issue_date") or "") or ""),
                    "valuation": None,
                    "identity_state": "UNRESOLVED",
                    "identity_method": "UNRESOLVED",
                    "raw_payload": row,
                }
            )
        parsed = stage_c_identity(parsed)
    return discovery_report(audit, parsed)
